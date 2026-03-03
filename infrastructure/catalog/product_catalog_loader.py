# infrastructure/catalog/product_catalog_loader.py
from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

from domain.models.bc3_classification_models import Bc3CatalogoItem


def _strip_accents(text: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn"
    )


def _norm_header(text: str) -> str:
    t = _strip_accents(text or "").strip().lower()
    t = re.sub(r"[^a-z0-9]+", "_", t)
    return t.strip("_")


def _split_tags(value: Optional[str]) -> List[str]:
    if not value:
        return []
    raw = str(value).strip()
    if not raw:
        return []
    parts = re.split(r"[;,|/\t]+", raw)
    out: List[str] = []
    for p in parts:
        p = p.strip()
        if p:
            out.append(p)
    return out


@dataclass(frozen=True)
class ProductCatalogColumnMap:
    code: str

    # Campos específicos “Ruesma”
    full_description: Optional[str]
    product_description: Optional[str]
    family_description: Optional[str]
    group_description: Optional[str]

    # Campos genéricos opcionales
    tags: Optional[str]
    extra_tag_cols: Tuple[str, ...]


class ProductCatalogLoader:
    """
    Carga catálogo interno desde CSV/XLSX.

    Estructura objetivo del Excel (según tu imagen):
    - "Codigo producto"           -> codigo
    - "Descripcion Completa"      -> descripcion_completa
    - "descripcion producto"      -> descripcion_producto
    - "descripcion familia"       -> descripcion_familia
    - "descripcion grupo"         -> descripcion_grupo
    """

    # --- mapeos robustos (normalizados) ---
    _CODE_KEYS = {
        "codigo",
        "code",
        "cod",
        "codigo_interno",
        "codigo_producto",
        "cod_producto",
        "id",
        "ref",
        "referencia",
    }

    _FULL_DESC_KEYS = {
        "descripcion_completa",
        "descripcioncompleta",
        "descripcion_total",
        "descripcion_larga",
        "descripcion",
        "description",
        "desc",
        "detalle",
        "texto",
        "concepto",
    }

    _PRODUCT_DESC_KEYS = {
        "descripcion_producto",
        "producto",
        "nombre",
        "name",
        "denominacion",
        "denominacion_producto",
    }

    _FAMILY_KEYS = {
        "descripcion_familia",
        "familia",
        "family",
        "familia_compra",
        "descripcion_familia_compra",
    }

    _GROUP_KEYS = {
        "descripcion_grupo",
        "grupo",
        "group",
        "tipo_coste",
        "tipo_de_coste",
        "tipo",
        "clase",
    }

    _TAGS_KEYS = {
        "tags",
        "etiquetas",
        "keywords",
        "palabras_clave",
    }

    # Columnas que aportan contexto y se pueden meter como tags extra
    _EXTRA_TAG_KEYS = {
        "familia",
        "grupo",
        "categoria",
        "subcategoria",
        "tipo",
        "clase",
        "subclase",
        "seccion",
        "unidad",
        "ud",
        "um",
        # importantes para tu excel:
        "descripcion_familia",
        "descripcion_grupo",
    }

    def load(
        self,
        *,
        path: Path,
        sheet_name: Optional[str] = None,
        encoding: str = "utf-8-sig",
    ) -> List[Bc3CatalogoItem]:
        if not path.exists():
            raise FileNotFoundError(f"Catálogo no encontrado: {path}")

        ext = path.suffix.lower().strip()
        if ext in {".xlsx", ".xlsm"}:
            rows, headers = self._read_xlsx(path, sheet_name=sheet_name)
            colmap = self._infer_columns(headers)
            return self._rows_to_items(rows, headers=headers, colmap=colmap)

        if ext in {".csv", ".txt"}:
            text = path.read_bytes().decode(encoding, errors="replace")
            rows, headers = self._read_csv_text(text)
            colmap = self._infer_columns(headers)
            return self._rows_to_items(rows, headers=headers, colmap=colmap)

        raise ValueError(f"Extensión de catálogo no soportada: {ext} (usa CSV o XLSX)")

    @staticmethod
    def _pick_best_delimiter(first_line: str) -> str:
        candidates = [";", ",", "\t", "|"]
        best = ";"
        best_count = -1
        for d in candidates:
            c = first_line.count(d)
            if c > best_count:
                best_count = c
                best = d
        return best

    def _read_csv_text(self, text: str) -> Tuple[List[Dict[str, str]], List[str]]:
        text = text.lstrip("\ufeff")
        lines = text.splitlines()
        if not lines:
            return [], []

        delimiter = self._pick_best_delimiter(lines[0])

        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        headers = reader.fieldnames or []
        rows: List[Dict[str, str]] = []
        for row in reader:
            cleaned: Dict[str, str] = {}
            for k, v in (row or {}).items():
                if k is None:
                    continue
                cleaned[str(k)] = "" if v is None else str(v)
            rows.append(cleaned)
        return rows, headers

    def _read_xlsx(
        self, path: Path, sheet_name: Optional[str]
    ) -> Tuple[List[Dict[str, str]], List[str]]:
        wb = load_workbook(path, read_only=True, data_only=True)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]

        rows_iter = ws.iter_rows(values_only=True)

        try:
            headers_row = next(rows_iter)
        except StopIteration:
            return [], []

        # Mantener posiciones para no desalinear si hay columnas vacías
        headers_all = [str(h).strip() if h is not None else "" for h in headers_row]
        headers = [h for h in headers_all if h]

        out_rows: List[Dict[str, str]] = []
        for r in rows_iter:
            if r is None:
                continue
            data: Dict[str, str] = {}
            for idx, header in enumerate(headers_all):
                if not header:
                    continue
                val = r[idx] if idx < len(r) else None
                data[header] = "" if val is None else str(val).strip()

            if all(not v for v in data.values()):
                continue
            out_rows.append(data)

        return out_rows, headers

    def _infer_columns(self, headers: List[str]) -> ProductCatalogColumnMap:
        if not headers:
            raise ValueError("El catálogo no tiene cabecera (headers vacíos).")

        norm_to_raw: Dict[str, str] = {}
        for h in headers:
            norm_to_raw[_norm_header(h)] = h

        def find_first(keys: Iterable[str]) -> Optional[str]:
            for k in keys:
                nk = _norm_header(k)
                if nk in norm_to_raw:
                    return norm_to_raw[nk]
            return None

        code_col = find_first(self._CODE_KEYS)
        if not code_col:
            raise ValueError(
                "No se detectó columna de código ('Codigo producto'). "
                f"Headers disponibles: {headers}"
            )

        full_col = find_first(self._FULL_DESC_KEYS)
        product_col = find_first(self._PRODUCT_DESC_KEYS)
        family_col = find_first(self._FAMILY_KEYS)
        group_col = find_first(self._GROUP_KEYS)

        tags_col = find_first(self._TAGS_KEYS)

        extra_tag_cols: List[str] = []
        for k in self._EXTRA_TAG_KEYS:
            c = find_first([k])
            if c and c not in {code_col, full_col, product_col, family_col, group_col, tags_col}:
                extra_tag_cols.append(c)

        return ProductCatalogColumnMap(
            code=code_col,
            full_description=full_col,
            product_description=product_col,
            family_description=family_col,
            group_description=group_col,
            tags=tags_col,
            extra_tag_cols=tuple(extra_tag_cols),
        )

    def _rows_to_items(
        self,
        rows: List[Dict[str, str]],
        *,
        headers: List[str],
        colmap: ProductCatalogColumnMap,
    ) -> List[Bc3CatalogoItem]:
        items: List[Bc3CatalogoItem] = []
        seen_codes: set[str] = set()

        for row in rows:
            code = (row.get(colmap.code) or "").strip()
            if not code:
                continue
            if code in seen_codes:
                continue
            seen_codes.add(code)

            full = (row.get(colmap.full_description) or "").strip() if colmap.full_description else ""
            product = (row.get(colmap.product_description) or "").strip() if colmap.product_description else ""
            family = (row.get(colmap.family_description) or "").strip() if colmap.family_description else ""
            group = (row.get(colmap.group_description) or "").strip() if colmap.group_description else ""

            # Tags: priorizamos grupo y familia porque son “nivel 1 y 2”
            tags: List[str] = []
            if group:
                tags.append(group)
            if family:
                tags.append(family)

            if colmap.tags:
                tags.extend(_split_tags(row.get(colmap.tags)))

            for c in colmap.extra_tag_cols:
                v = (row.get(c) or "").strip()
                if v:
                    tags.append(v)

            # Dedup tags
            tags_norm: List[str] = []
            seen_t = set()
            for t in tags:
                tt = t.strip()
                if not tt:
                    continue
                key = tt.lower()
                if key in seen_t:
                    continue
                seen_t.add(key)
                tags_norm.append(tt)

            # Para scoring legacy:
            # - nombre => descripcion_producto
            # - descripcion => descripcion_completa (si existe) o composición
            name = product or None
            desc = full or None
            if not desc:
                # fallback: monta una descripción consistente
                parts = [p for p in [group, family, product] if p]
                desc = ", ".join(parts) if parts else None

            items.append(
                Bc3CatalogoItem(
                    codigo=code,
                    nombre=name,
                    descripcion=desc,
                    tags=tags_norm,
                    descripcion_completa=full or None,
                    descripcion_producto=product or None,
                    descripcion_familia=family or None,
                    descripcion_grupo=group or None,
                )
            )

        if not items:
            raise ValueError(
                "El catálogo se cargó pero no produjo ningún item (0 códigos válidos). "
                "Revisa que exista la columna 'Codigo producto' y tenga valores."
            )

        return items